import json
import openpyxl
from tkinter import filedialog
import tkinter as tk
import pandas as pd
from github import Github

# Create a file dialog to select an Excel or CSV file
root = tk.Tk()
root.withdraw()

Fname = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"),
                                              ("CSV files", "*.csv"),
                                               ("All files", "*.*")))

# If the selected file is a CSV file, convert it to an Excel file
# and update the filename
if Fname[-3:] == "csv":
    read_file = pd.read_csv (Fname,encoding = "ISO-8859-1")
    read_file.to_excel (Fname[:-3] + "xlsx", index = None, header=True)
    Fname = Fname[:-3] + "xlsx"

# Initialize dictionaries and lists for storing information
Rules = {}
Total_Tactics = []
Total_Techniques = []

# Define column names to be used for extracting data from the Excel sheet
Rule_ColName = 'Rule Name'
Mitre_Tactic_ColName = 'MITRE Tactic'
Mitre_Technique_ColName = 'MITRE Technique'

# Specify names for output files
Json_file = "MITRE_Matrix.json"
html_file = "Mitre Navigator.html"

# Github Access Token for authentication
Github_Access_Token = ""

# Name of the Github repository where the files will be uploaded
Repo_name = 'Automation-Scripts'
git_prefix = 'MitreATT&CK/Mapping Rules to navigator/'

# Load the Excel file and select the active sheet
wb = openpyxl.load_workbook(Fname)
ws = wb.active

# Create a dictionary of column names
ColNames = {}
for COL in ws.iter_cols(1, 12):
    ColNames[COL[0].value] = COL[0].column_letter

#print(ColNames)
Rule = ColNames[Rule_ColName]
Mitre_Tactic = ColNames[Mitre_Tactic_ColName]
Mitre_Technique = ColNames[Mitre_Technique_ColName]

print("[+] The Rules sheet was opened successfully")

# Extract techniques from the Rules sheet of the Excel file
for row in ws[2:ws.max_row]:
    Rule, Mitre_Tactic, Mitre_Technique = ColNames[Rule_ColName], ColNames[Mitre_Tactic_ColName], ColNames[Mitre_Technique_ColName]
    Rule += str(row[0].row)
    Mitre_Tactic += str(row[0].row)
    Mitre_Technique += str(row[0].row)
    if ws[Mitre_Tactic].value == ('#N/A' or '' or ' ' or 'None'):
        continue
    elif ws[Mitre_Tactic].value is None:
        continue
    elif ws[Rule].value not in Rules:
        Rules[ws[Rule].value] = {}
        #print(ws[Mitre_Tactic].value,ws[Mitre_Technique].value)
        Rules[ws[Rule].value]["Tactic"] = ws[Mitre_Tactic].value.replace('\n', '').replace(r"\n", '').split(",")
        Rules[ws[Rule].value]["Technique"] = ws[Mitre_Technique].value.replace('\n', '').replace(r"\n", '').split(",")

# Get sub-techniques and add them to the list of techniques
for rule in Rules.keys():
    Total_Techniques += Rules[rule]["Technique"]
    Total_Tactics += Rules[rule]["Tactic"]

te=[x[0:5] for x in Total_Techniques]

Total_Techniques += te

Total_Techniques = list(set(Total_Techniques))

print("[+] The Mitre Techniques was extracted successfully")

#print(Total_Tactics)
#print(Total_Techniques)

# Generate MITRE Layer
Layer_Template = {
    "name": "Rules Coverage",
	"versions": {
		"attack": "13",
		"navigator": "4.8.2",
		"layer": "4.4"
	},
	"domain": "enterprise-attack",
	"description": "Techniques Covered by Rules",
	"filters": {
		"platforms": [
			"Linux",
			"macOS",
			"Windows",
			"PRE",
			"Containers",
			"Network",
			"Office 365",
			"SaaS",
			"Google Workspace",
			"IaaS",
			"Azure AD"
		]
	},
	"sorting": 0,
	"layout": {
		"layout": "side",
		"aggregateFunction": "average",
		"showID": False,
		"showName": True,
		"showAggregateScores": False,
		"countUnscored": False
	},
	"hideDisabled": False,
    "techniques":
        [{"techniqueID": technique, "color": "#3182bd"} for technique in Total_Techniques]
    ,
    "gradient": {
		"colors": [
			"#ff6666ff",
			"#ffe766ff",
			"#8ec843ff"
		],
        "minValue": 0,
        "maxValue": 1
    },
    "legendItems": [],
	"metadata": [],
	"links": [],
	"showTacticRowBackground": False,
	"tacticRowBackground": "#dddddd",
	"selectTechniquesAcrossTactics": True,
	"selectSubtechniquesWithParent": True
}

# Converts a dictionary to a JSON-formatted string
json_data = json.dumps(Layer_Template)

# Creates or overwrites a JSON file with the JSON-formatted string
# 'sort_keys' sorts the keys alphabetically
# 'indent' specifies the number of spaces for indentation
with open(Json_file, "w") as file:
    json.dump(Layer_Template, file, sort_keys=True, indent=4)

print(f"[+] The MITRE matrix json file '{Json_file}' was created successfully")

# Initializes an instance of the Github class using a GitHub access token
g = Github(Github_Access_Token)

# Gets the repository specified by the 'Repo_name' variable
# Gets a list of all files in the repository
repo = g.get_user().get_repo(Repo_name)
all_files = []
contents = repo.get_contents("")
for content_file in contents:
    file = content_file
    all_files.append(str(file).replace('ContentFile(path="','').replace('")',''))

# Reads the contents of the JSON file and stores them in a variable
with open(Json_file, 'r') as file:
    content = file.read()

# Gets a list of all branches in the repository
# Selects the first branch in the list and sets it as the active branch
l = list(repo.get_branches())
branch = str(l[0]).replace('Branch(name="','').replace('")','')

# Constructs the path to the JSON file on GitHub
git_file = git_prefix + Json_file

# If the JSON file already exists on GitHub, updates it with the new content
# Otherwise, creates a new file with the specified content
if git_file in all_files:
    contents = repo.get_contents(git_file)
    repo.update_file(contents.path, "committing files", content, contents.sha, branch=branch)
    print("[+] The MITRE matrix json file '" + git_file + "' UPDATED on GitHub Repo successfully")
else:
    repo.create_file(git_file, "committing files", content, branch=branch)
    print("[+] The MITRE matrix json file '" + git_file + "' CREATED on GitHub Repo successfully")

# Gets the download URL for the JSON file on GitHub and constructs a URL for the MITRE Navigator
contents = repo.get_contents(Json_file)
raw_url = contents.download_url

Mitre_url = f"https://mitre-attack.github.io/attack-navigator/#layerURL={raw_url}&tabs=false&selecting_techniques=false"

print("[+] The MITRE matrix json IFrame URL was generated successfully")

# Creates or overwrites an HTML file with the code for an IFrame that displays the MITRE Navigator
with open(html_file, "w") as html:
    # the html code which will go in the file 'Mitre Navigator.html'
	html_template = f"""<html>
	<head>
	<title>Mitre Navigator</title>
	</head>
	<body>
	<iframe src={Mitre_url} width="100%" height="100%"></iframe>
	
	</body>
	</html>
	"""
	
	# writing the code into the file
	html.write(html_template)

print(f"[+] The MITRE Navigator HTML file '{html_file}' was created successfully")